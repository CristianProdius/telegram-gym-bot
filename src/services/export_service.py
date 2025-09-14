import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from src.models import Workout, WorkoutExercise, WorkoutSet, Exercise, User, Routine, ProgressRecord
from src.database.connection import get_session

logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting workout data in various formats"""

    async def export_to_excel(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Export workout data to Excel format"""
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()

        async with get_session() as session:
            # Get workouts
            stmt = (
                select(Workout)
                .where(
                    Workout.user_id == user_id,
                    Workout.date >= start_date,
                    Workout.date <= end_date
                )
                .options(
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.exercise),
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.sets)
                )
                .order_by(Workout.date.desc())
            )

            result = await session.execute(stmt)
            workouts = result.scalars().unique().all()

            # Create workbook
            wb = Workbook()

            # Workout Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Workout Summary"
            self._create_summary_sheet(ws_summary, workouts)

            # Detailed Workouts Sheet
            ws_detail = wb.create_sheet("Detailed Workouts")
            self._create_detail_sheet(ws_detail, workouts)

            # Statistics Sheet
            ws_stats = wb.create_sheet("Statistics")
            await self._create_stats_sheet(ws_stats, user_id, session)

            # Personal Records Sheet
            ws_pr = wb.create_sheet("Personal Records")
            await self._create_pr_sheet(ws_pr, user_id, session)

            # Save to bytes
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            return excel_file.getvalue()

    def _create_summary_sheet(self, ws, workouts):
        """Create summary sheet in Excel"""
        # Headers
        headers = ["Date", "Exercise", "Sets", "Total Reps", "Total Volume", "Max Weight", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data
        row = 2
        for workout in workouts:
            for workout_exercise in workout.exercises:
                exercise = workout_exercise.exercise
                total_reps = sum(s.reps for s in workout_exercise.sets)
                total_volume = sum(s.reps * s.weight for s in workout_exercise.sets)
                max_weight = max((s.weight for s in workout_exercise.sets), default=0)

                ws.cell(row=row, column=1, value=workout.date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=exercise.name)
                ws.cell(row=row, column=3, value=len(workout_exercise.sets))
                ws.cell(row=row, column=4, value=total_reps)
                ws.cell(row=row, column=5, value=round(total_volume, 2))
                ws.cell(row=row, column=6, value=max_weight)
                ws.cell(row=row, column=7, value=workout_exercise.notes or "")

                row += 1

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_detail_sheet(self, ws, workouts):
        """Create detailed workout sheet in Excel"""
        headers = ["Date", "Exercise", "Set", "Reps", "Weight", "RPE", "Rest (sec)", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        row = 2
        for workout in workouts:
            for workout_exercise in workout.exercises:
                exercise = workout_exercise.exercise
                for workout_set in workout_exercise.sets:
                    ws.cell(row=row, column=1, value=workout.date.strftime("%Y-%m-%d"))
                    ws.cell(row=row, column=2, value=exercise.name)
                    ws.cell(row=row, column=3, value=workout_set.set_number)
                    ws.cell(row=row, column=4, value=workout_set.reps)
                    ws.cell(row=row, column=5, value=workout_set.weight)
                    ws.cell(row=row, column=6, value=workout_set.rpe or "")
                    ws.cell(row=row, column=7, value=workout_set.rest_seconds or "")
                    ws.cell(row=row, column=8, value=workout_set.notes or "")
                    row += 1

        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12

    async def _create_stats_sheet(self, ws, user_id, session):
        """Create statistics sheet in Excel"""
        from src.services.workout_service import WorkoutService
        from src.services.analytics_service import WorkoutAnalytics

        workout_service = WorkoutService()
        analytics = WorkoutAnalytics()

        # Get statistics
        stats = await workout_service.get_user_statistics(user_id)
        volume_data = await analytics.calculate_volume_progression(user_id, weeks=12)

        # Add headers
        ws.cell(row=1, column=1, value="Statistic").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Value").font = Font(bold=True)

        # Add statistics
        stats_data = [
            ("Total Workouts", stats["total_workouts"]),
            ("This Week", stats["week_workouts"]),
            ("Total Volume (kg)", stats["total_volume"]),
            ("Favorite Exercise", stats["favorite_exercise"]),
            ("Weekly Trend", f"{volume_data.get('trend', 0):.2f} kg/week"),
        ]

        for i, (label, value) in enumerate(stats_data, 2):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20

    async def _create_pr_sheet(self, ws, user_id, session):
        """Create personal records sheet in Excel"""
        from src.services.analytics_service import WorkoutAnalytics

        analytics = WorkoutAnalytics()
        records = await analytics.get_personal_records(user_id)

        # Headers
        headers = ["Exercise", "Record Type", "Value", "Date Achieved"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data
        for i, record in enumerate(records, 2):
            ws.cell(row=i, column=1, value=record["exercise"])
            ws.cell(row=i, column=2, value=record["type"])
            ws.cell(row=i, column=3, value=record["value"])
            ws.cell(row=i, column=4, value=record["date"].strftime("%Y-%m-%d"))

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15

    async def export_to_pdf(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Export workout data to PDF format"""
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()

        async with get_session() as session:
            # Get user info
            stmt = select(User).where(User.id == user_id)
            result = await session.execute(stmt)
            user = result.scalar_one_or_none()

            # Get workouts
            stmt = (
                select(Workout)
                .where(
                    Workout.user_id == user_id,
                    Workout.date >= start_date,
                    Workout.date <= end_date
                )
                .options(
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.exercise),
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.sets)
                )
                .order_by(Workout.date.desc())
            )

            result = await session.execute(stmt)
            workouts = result.scalars().unique().all()

            # Create PDF
            pdf_file = BytesIO()
            doc = SimpleDocTemplate(pdf_file, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=1  # Center
            )

            elements.append(Paragraph("Workout Report", title_style))
            elements.append(Spacer(1, 20))

            # User info
            if user:
                user_info = f"<b>User:</b> {user.username or 'N/A'}<br/>"
                user_info += f"<b>Period:</b> {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}<br/>"
                user_info += f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                elements.append(Paragraph(user_info, styles['Normal']))
                elements.append(Spacer(1, 20))

            # Statistics summary
            from src.services.workout_service import WorkoutService
            workout_service = WorkoutService()
            stats = await workout_service.get_user_statistics(user_id)

            stats_data = [
                ['Statistic', 'Value'],
                ['Total Workouts', str(stats['total_workouts'])],
                ['This Week', str(stats['week_workouts'])],
                ['Total Volume', f"{stats['total_volume']} kg"],
                ['Favorite Exercise', stats['favorite_exercise']]
            ]

            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(Paragraph("<b>Statistics Summary</b>", styles['Heading2']))
            elements.append(stats_table)
            elements.append(PageBreak())

            # Workout details
            elements.append(Paragraph("<b>Workout Details</b>", styles['Heading2']))
            elements.append(Spacer(1, 12))

            for workout in workouts[:20]:  # Limit to 20 most recent workouts
                # Workout date header
                date_str = workout.date.strftime("%A, %B %d, %Y")
                elements.append(Paragraph(f"<b>{date_str}</b>", styles['Heading3']))

                # Exercise data
                for workout_exercise in workout.exercises:
                    exercise = workout_exercise.exercise
                    exercise_data = []

                    # Headers
                    exercise_data.append([
                        exercise.name,
                        'Set',
                        'Reps',
                        'Weight (kg)',
                        'Volume'
                    ])

                    # Sets data
                    for workout_set in workout_exercise.sets:
                        volume = workout_set.reps * workout_set.weight
                        exercise_data.append([
                            '',
                            str(workout_set.set_number),
                            str(workout_set.reps),
                            str(workout_set.weight),
                            f"{volume:.1f}"
                        ])

                    # Create table
                    exercise_table = Table(exercise_data, colWidths=[2*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1*inch])
                    exercise_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))

                    elements.append(exercise_table)
                    elements.append(Spacer(1, 10))

                elements.append(Spacer(1, 20))

            # Build PDF
            doc.build(elements)
            pdf_file.seek(0)

            return pdf_file.getvalue()

    async def export_to_csv(
        self,
        user_id: int,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Export workout data to CSV format"""
        if not start_date:
            start_date = datetime.now() - timedelta(days=30)
        if not end_date:
            end_date = datetime.now()

        async with get_session() as session:
            stmt = (
                select(Workout)
                .where(
                    Workout.user_id == user_id,
                    Workout.date >= start_date,
                    Workout.date <= end_date
                )
                .options(
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.exercise),
                    selectinload(Workout.exercises)
                    .selectinload(WorkoutExercise.sets)
                )
                .order_by(Workout.date.desc())
            )

            result = await session.execute(stmt)
            workouts = result.scalars().unique().all()

            # Create data for CSV
            data = []
            for workout in workouts:
                for workout_exercise in workout.exercises:
                    exercise = workout_exercise.exercise
                    for workout_set in workout_exercise.sets:
                        data.append({
                            'Date': workout.date.strftime('%Y-%m-%d'),
                            'Exercise': exercise.name,
                            'Category': exercise.category,
                            'Muscle Group': exercise.muscle_group,
                            'Set': workout_set.set_number,
                            'Reps': workout_set.reps,
                            'Weight (kg)': workout_set.weight,
                            'Volume (kg)': workout_set.reps * workout_set.weight,
                            'RPE': workout_set.rpe or '',
                            'Rest (sec)': workout_set.rest_seconds or '',
                            'Notes': workout_set.notes or ''
                        })

            # Create DataFrame and export to CSV
            df = pd.DataFrame(data)
            csv_file = BytesIO()
            df.to_csv(csv_file, index=False, encoding='utf-8')
            csv_file.seek(0)

            return csv_file.getvalue()

    async def export_routine_to_pdf(self, routine_id: int) -> bytes:
        """Export a workout routine to PDF"""
        from src.services.routine_service import RoutineService

        routine_service = RoutineService()
        routine_details = await routine_service.get_routine_details(routine_id)

        if not routine_details:
            raise ValueError("Routine not found")

        # Create PDF
        pdf_file = BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = f"Workout Routine: {routine_details['name']}"
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))

        # Description
        if routine_details['description']:
            elements.append(Paragraph(routine_details['description'], styles['Normal']))
            elements.append(Spacer(1, 12))

        # Metadata
        meta_info = f"<b>Category:</b> {routine_details['category'] or 'N/A'}<br/>"
        meta_info += f"<b>Difficulty:</b> {routine_details['difficulty']}<br/>"
        meta_info += f"<b>Total Exercises:</b> {routine_details['total_exercises']}"
        elements.append(Paragraph(meta_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Exercises by day
        for day, exercises in routine_details['exercises_by_day'].items():
            elements.append(Paragraph(f"<b>Day {day}</b>", styles['Heading2']))

            exercise_data = [['Exercise', 'Sets', 'Reps', 'Rest']]
            for ex in exercises:
                exercise_data.append([
                    ex['exercise'],
                    str(ex['target_sets'] or '-'),
                    ex['target_reps'] or '-',
                    f"{ex['rest_seconds']}s"
                ])

            table = Table(exercise_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)

        return pdf_file.getvalue()